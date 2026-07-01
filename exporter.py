"""
exporter.py

Exports ranked candidates to an Excel (.xlsx) file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "output/ranked_candidates.xlsx"


def export_rankings(ranked_candidates):
    """
    Export ranked candidates to an Excel (.xlsx) file.
    """

    os.makedirs("output", exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ranked Candidates"

    headers = [
        "Rank",
        "Candidate ID",
        "Current Title",
        "Current Company",
        "Experience (Years)",
        "Experience Score",
        "Skills Score",
        "Career Score",
        "Profile Score",
        "Signal Score",
        "Total Score",
        "Explanation"
    ]

    worksheet.append(headers)

    # Make header bold
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Write candidate data
    for rank, result in enumerate(ranked_candidates, start=1):

        candidate = result["candidate"]
        profile = candidate.get("profile", {})

        explanation = " | ".join(result["explanation"])

        worksheet.append([
            rank,
            candidate.get("candidate_id"),
            profile.get("current_title"),
            profile.get("current_company"),
            profile.get("years_of_experience"),
            result["experience"],
            result["skills"],
            result["career"],
            result["profile"],
            result["signals"],
            result["total_score"],
            explanation
        ])

    # Auto-adjust column widths
    for column_cells in worksheet.columns:

        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:

            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 60)

        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = adjusted_width

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Enable filters
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(OUTPUT_FILE)